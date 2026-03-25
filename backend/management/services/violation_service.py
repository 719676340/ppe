# backend/management/services/violation_service.py
from sqlalchemy.orm import Session
from management.database import Violation, Camera, DetectionZone
from management.schemas.violation import ViolationFilter
from typing import List, Optional, Dict
from datetime import datetime
import logging
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
import io

logger = logging.getLogger(__name__)

class ViolationService:
    @staticmethod
    def get_violations(
        db: Session,
        camera_id: Optional[int] = None,
        zone_id: Optional[int] = None,
        start_time: Optional[datetime] = None,
        end_time: Optional[datetime] = None,
        is_processed: Optional[bool] = None,
        skip: int = 0,
        limit: int = 20
    ) -> tuple[List[Dict], int]:
        """获取违规记录列表"""
        # 使用 join 来获取关联的 camera 和 zone 信息
        query = db.query(
            Violation.id,
            Violation.camera_id,
            Violation.zone_id,
            Violation.violation_time,
            Violation.image_path,
            Violation.is_processed,
            Violation.remark,
            Violation.created_at,
            Camera.name.label('camera_name'),
            DetectionZone.name.label('zone_name')
        ).join(
            Camera, Violation.camera_id == Camera.id
        ).outerjoin(
            DetectionZone, Violation.zone_id == DetectionZone.id
        )

        if camera_id is not None:
            query = query.filter(Violation.camera_id == camera_id)
        if zone_id is not None:
            query = query.filter(Violation.zone_id == zone_id)
        if start_time is not None:
            query = query.filter(Violation.violation_time >= start_time)
        if end_time is not None:
            query = query.filter(Violation.violation_time <= end_time)
        if is_processed is not None:
            query = query.filter(Violation.is_processed == is_processed)

        total = query.count()
        results = query.order_by(Violation.violation_time.desc()).offset(skip).limit(limit).all()

        # 转换为字典列表
        violations = []
        for row in results:
            violations.append({
                'id': row.id,
                'camera_id': row.camera_id,
                'zone_id': row.zone_id,
                'violation_time': row.violation_time.isoformat() if row.violation_time else None,
                'image_path': row.image_path,
                'is_processed': row.is_processed,
                'remark': row.remark,
                'created_at': row.created_at.isoformat() if row.created_at else None,
                'camera_name': row.camera_name,
                'zone_name': row.zone_name
            })

        return violations, total

    @staticmethod
    def get_violation(db: Session, violation_id: int) -> Optional[Dict]:
        """获取单个违规记录"""
        result = db.query(
            Violation.id,
            Violation.camera_id,
            Violation.zone_id,
            Violation.violation_time,
            Violation.image_path,
            Violation.is_processed,
            Violation.remark,
            Violation.created_at,
            Camera.name.label('camera_name'),
            DetectionZone.name.label('zone_name')
        ).join(
            Camera, Violation.camera_id == Camera.id
        ).outerjoin(
            DetectionZone, Violation.zone_id == DetectionZone.id
        ).filter(
            Violation.id == violation_id
        ).first()

        if not result:
            return None

        return {
            'id': result.id,
            'camera_id': result.camera_id,
            'zone_id': result.zone_id,
            'violation_time': result.violation_time.isoformat() if result.violation_time else None,
            'image_path': result.image_path,
            'is_processed': result.is_processed,
            'remark': result.remark,
            'created_at': result.created_at.isoformat() if result.created_at else None,
            'camera_name': result.camera_name,
            'zone_name': result.zone_name
        }

    @staticmethod
    def update_violation_remark(db: Session, violation_id: int, remark: str) -> Optional[Violation]:
        """更新违规备注"""
        db_violation = db.query(Violation).filter(Violation.id == violation_id).first()
        if not db_violation:
            return None

        db_violation.remark = remark
        db.commit()
        db.refresh(db_violation)
        return db_violation

    @staticmethod
    def mark_as_processed(db: Session, violation_id: int) -> Optional[Violation]:
        """标记为已处理"""
        db_violation = db.query(Violation).filter(Violation.id == violation_id).first()
        if not db_violation:
            return None

        db_violation.is_processed = True
        db.commit()
        db.refresh(db_violation)
        return db_violation

    @staticmethod
    def delete_violation(db: Session, violation_id: int) -> bool:
        """删除违规记录"""
        db_violation = db.query(Violation).filter(Violation.id == violation_id).first()
        if not db_violation:
            return False

        db.delete(db_violation)
        db.commit()
        return True

    @staticmethod
    def batch_mark_processed(db: Session, violation_ids: List[int]) -> int:
        """批量标记为已处理"""
        count = db.query(Violation).filter(Violation.id.in_(violation_ids)).update(
            {"is_processed": True}
        )
        db.commit()
        return count

    @staticmethod
    def _prepare_image_for_excel(image_path: Path, max_width: int = 150, max_height: int = 100) -> Optional[XLImage]:
        """准备图片用于嵌入Excel"""
        try:
            xl_img = XLImage(str(image_path))

            # 获取原始尺寸并计算缩放
            orig_width, orig_height = xl_img.width, xl_img.height
            scale_w = max_width / orig_width if orig_width > max_width else 1
            scale_h = max_height / orig_height if orig_height > max_height else 1
            scale = min(scale_w, scale_h)

            xl_img.width = int(orig_width * scale)
            xl_img.height = int(orig_height * scale)

            return xl_img

        except Exception as e:
            logger.error(f"处理图片失败 {image_path}: {e}", exc_info=True)
            return None

    @staticmethod
    def export_to_excel(
        db: Session,
        camera_id: Optional[int] = None,
        zone_id: Optional[int] = None,
        start_time: Optional[datetime] = None,
        end_time: Optional[datetime] = None,
        is_processed: Optional[bool] = None,
        base_dir: Path = None
    ) -> bytes:
        """导出违规记录到Excel文件，包含截图"""
        query = db.query(
            Violation.id,
            Violation.camera_id,
            Violation.zone_id,
            Violation.violation_time,
            Violation.image_path,
            Violation.is_processed,
            Violation.remark,
            Camera.name.label('camera_name'),
            DetectionZone.name.label('zone_name')
        ).join(
            Camera, Violation.camera_id == Camera.id
        ).outerjoin(
            DetectionZone, Violation.zone_id == DetectionZone.id
        )

        if camera_id is not None:
            query = query.filter(Violation.camera_id == camera_id)
        if zone_id is not None:
            query = query.filter(Violation.zone_id == zone_id)
        if start_time is not None:
            query = query.filter(Violation.violation_time >= start_time)
        if end_time is not None:
            query = query.filter(Violation.violation_time <= end_time)
        if is_processed is not None:
            query = query.filter(Violation.is_processed == is_processed)

        results = query.order_by(Violation.violation_time.desc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "违规记录"

        headers = ["ID", "摄像头", "区域", "违规时间", "处理状态", "备注", "截图"]
        header_col_widths = [8, 20, 20, 20, 12, 30, 25]

        border_style = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_alignment = Alignment(horizontal='center', vertical='center')

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border_style
            ws.column_dimensions[get_column_letter(col_num)].width = header_col_widths[col_num - 1]

        ws.row_dimensions[1].height = 25

        img_max_width = 150
        img_max_height = 100

        images_count = 0
        for row_num, row in enumerate(results, 2):
            ws.row_dimensions[row_num].height = img_max_height + 10

            data = [
                row.id,
                row.camera_name or '',
                row.zone_name or '',
                row.violation_time.strftime('%Y-%m-%d %H:%M:%S') if row.violation_time else '',
                '已处理' if row.is_processed else '未处理',
                row.remark or '',
                ''
            ]

            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border_style
                cell.alignment = Alignment(horizontal='left', vertical='center')
                if col_num in [1, 5]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            if row.image_path:
                image_path = None
                if base_dir:
                    image_path = base_dir / row.image_path.lstrip('/')
                else:
                    image_path = Path(row.image_path)

                logger.info(f"处理图片: {image_path}, 存在: {image_path.exists()}")

                if image_path.exists():
                    xl_img = ViolationService._prepare_image_for_excel(image_path, img_max_width, img_max_height)
                    if xl_img:
                        img_cell = f'G{row_num}'
                        ws.add_image(xl_img, img_cell)
                        images_count += 1
                        logger.info(f"成功嵌入图片 {row_num}: {image_path}")
                    else:
                        logger.warning(f"跳过无效图片: {image_path}")

        logger.info(f"Excel导出完成，共 {len(results)} 条记录，成功嵌入 {images_count} 张图片")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()
